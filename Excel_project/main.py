import openpyxl


inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_sup = {}
total_value_per_sup = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory =  product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_n = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)


    # calculation for number of products for supplier
    if supplier_name in products_per_sup:
       current_num_products = products_per_sup.get(supplier_name)
       products_per_sup[supplier_name] = current_num_products + 1

    else:
        products_per_sup[supplier_name] = 1

    #calculation of total value of inventory per supplier
    if supplier_name in total_value_per_sup:
        current_total_value = total_value_per_sup.get(supplier_name)
        total_value_per_sup[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_sup[supplier_name] = inventory * price

    #colculation  for  prod under 10
    if inventory < 10:
        products_under_10_inv[int(product_n)] = int(inventory)

    #add a inventory total value price
    inventory_price.value = inventory * price

print(products_per_sup)
print(total_value_per_sup)
print(products_under_10_inv)

inv_file.save("inventory_with_total_value.xlsx")
